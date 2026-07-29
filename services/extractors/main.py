#!/usr/bin/env python3
"""CBSL Money Market production slice: fetch → validate → ingest pending observations.

Locked reports (default): 5206, 1059, 1064
Series: Call/Repo, SDF/SLF/liquidity, Term Repo/RR auctions
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import httpx
from bs4 import BeautifulSoup

API_URL = os.environ.get("API_URL", "http://localhost:4000")
RAW_DIR = Path(__file__).resolve().parent / "data" / "raw"
FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
RAW_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SOURCES = {
    "5206": {
        "entry": "https://www.cbsl.lk/eResearch/MoneyMarketRatesDefault.aspx?ReportId=5206",
        "title": "Money Market Summary",
    },
    "1059": {
        "entry": "https://www.cbsl.lk/eResearch/MoneyMarketRatesDefault.aspx?ReportId=1059",
        "title": "Daily Operations",
    },
    "6169": {
        "entry": "https://www.cbsl.lk/eResearch/MoneyMarketRatesDefault.aspx?ReportId=6169",
        "title": "Government Securities",
    },
    "1064": {
        "entry": "https://www.cbsl.lk/eResearch/MoneyMarketRatesDefault.aspx?ReportId=1064",
        "title": "OMO Term Repo / Reverse Repo",
    },
    "6277": {
        "entry": "https://www.cbsl.lk/eResearch/MoneyMarketRatesDefault.aspx?ReportId=6277",
        "title": "Commercial Bank Lending and Deposit Rates",
    },
    "policy-rates": {
        "entry": "https://www.cbsl.gov.lk/en/rates-and-indicators/policy-rates",
        "title": "Policy Rates (OPR / SRR)",
        "kind": "xlsx_history",
        "xlsx_fallback": (
            "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/about/"
            "historical_policy_interest_rates.xlsx"
        ),
    },
    "consumer-price-inflation": {
        "entry": "https://www.cbsl.gov.lk/en/measures-of-consumer-price-inflation",
        "title": "Consumer Price Inflation (CCPI / NCPI)",
        "kind": "html_iframe_table",
        # Table lives in a custom iframe, not the Drupal page body.
        "table": "https://www.cbsl.gov.lk/cbsl_custom/inflation/inflationwindow.php",
    },
    "daily-economic-indicators": {
        "entry": (
            "https://www.cbsl.gov.lk/en/statistics/economic-indicators/daily-indicators"
        ),
        "title": "Daily Economic Indicators",
        "kind": "pdf_listing",
    },
    "weekly-economic-indicators": {
        "entry": (
            "https://www.cbsl.gov.lk/en/statistics/economic-indicators/weekly-indicators"
        ),
        "title": "Weekly Economic Indicators",
        "kind": "pdf_listing",
    },
    "monthly-economic-indicators": {
        "entry": (
            "https://www.cbsl.gov.lk/en/statistics/economic-indicators/monthly-indicators"
        ),
        "title": "Monthly Economic Indicators",
        "kind": "pdf_listing",
    },
    "external-sector-performance": {
        "entry": (
            "https://www.cbsl.gov.lk/en/press/press-releases/external-sector-performance"
        ),
        "title": "External Sector Performance",
        "kind": "pdf_listing",
    },
}

# Full-history sources — one fetch, ignore date-chunk windows.
FULL_HISTORY_REPORTS = frozenset({"policy-rates", "consumer-price-inflation"})

# PDF listing scrapers — discover report PDFs, filter by date window.
PDF_LISTING_REPORTS = frozenset(
    {
        "daily-economic-indicators",
        "weekly-economic-indicators",
        "monthly-economic-indicators",
        "external-sector-performance",
    }
)

# Value bounds for production validation
BOUNDS: dict[str, tuple[float, float]] = {
    "sl.mm.call.wa_yield": (0.0, 40.0),
    "sl.mm.call.min_rate": (0.0, 40.0),
    "sl.mm.call.max_rate": (0.0, 40.0),
    "sl.mm.call.volume": (0.0, 5000.0),
    "sl.mm.repo.wa_yield": (0.0, 40.0),
    "sl.mm.repo.min_rate": (0.0, 40.0),
    "sl.mm.repo.max_rate": (0.0, 40.0),
    "sl.mm.repo.volume": (0.0, 5000.0),
    "sl.mm.omo.offer_repo": (0.0, 5000.0),
    "sl.mm.omo.offer_reverse_repo": (0.0, 5000.0),
    "sl.mm.omo.received": (0.0, 5000.0),
    "sl.mm.omo.accepted": (0.0, 5000.0),
    "sl.mm.omo.min_rate": (0.0, 40.0),
    "sl.mm.omo.max_rate": (0.0, 40.0),
    "sl.mm.omo.wa_yield": (0.0, 40.0),
    "sl.mm.sdf.rate": (0.0, 40.0),
    "sl.mm.sdf.volume": (0.0, 5000.0),
    "sl.mm.slf.rate": (0.0, 40.0),
    "sl.mm.slf.volume": (0.0, 5000.0),
    "sl.mm.overnight_liquidity": (-5000.0, 5000.0),
    "sl.mm.cbsl.gov_holdings": (0.0, 20000.0),
    "sl.mm.tbill.91d": (0.0, 40.0),
    "sl.mm.tbill.182d": (0.0, 40.0),
    "sl.mm.tbill.364d": (0.0, 40.0),
    "sl.mm.term_repo.offer_repo": (0.0, 5000.0),
    "sl.mm.term_repo.offer_reverse_repo": (0.0, 5000.0),
    "sl.mm.term_repo.received": (0.0, 5000.0),
    "sl.mm.term_repo.accepted": (0.0, 5000.0),
    "sl.mm.term_repo.min_rate": (0.0, 40.0),
    "sl.mm.term_repo.max_rate": (0.0, 40.0),
    "sl.mm.term_repo.wa_yield": (0.0, 40.0),
    "sl.mm.term_repo.tenure": (1.0, 365.0),
    "sl.mm.term_repo.settlement_date": (20200101.0, 20401231.0),
    "sl.mm.term_repo.maturity_date": (20200101.0, 20401231.0),
    "sl.mm.term_repo.side": (1.0, 2.0),
    "sl.mm.awpr": (0.0, 40.0),
    "sl.mm.awpr.monthly": (0.0, 40.0),
    "sl.mm.awpr.6m": (0.0, 40.0),
    "sl.mm.awlr": (0.0, 40.0),
    "sl.mm.awdr": (0.0, 40.0),
    "sl.mm.awdr.6m": (0.0, 40.0),
    "sl.mm.awfdr": (0.0, 40.0),
    "sl.mm.awsr": (0.0, 40.0),
    "sl.mm.awndr": (0.0, 40.0),
    "sl.mm.awnlr": (0.0, 40.0),
    "sl.mm.opr": (0.0, 40.0),
    "sl.mm.srr": (0.0, 40.0),
    # YoY % — 2022 peak was ~70%; allow deep negatives for 2024–25 deflation.
    "sl.ei.ccpi.headline_yoy": (-30.0, 100.0),
    "sl.ei.ccpi.core_yoy": (-30.0, 100.0),
    "sl.ei.ncpi.headline_yoy": (-30.0, 100.0),
    "sl.ei.ncpi.core_yoy": (-30.0, 100.0),
    "sl.ei.ccpi.headline_mom": (-30.0, 30.0),
    "sl.ei.ncpi.headline_mom": (-30.0, 30.0),
    "sl.ei.ncpi.food_yoy": (-50.0, 100.0),
    "sl.ei.ncpi.nonfood_yoy": (-50.0, 100.0),
    # Daily / Weekly / Monthly Economic Indicators (PDF)
    "sl.fx.usd.tt_buy": (50.0, 1000.0),
    "sl.fx.usd.tt_sell": (50.0, 1000.0),
    "sl.fx.usd.spot": (50.0, 1000.0),
    "sl.fx.gbp.tt_buy": (50.0, 2000.0),
    "sl.fx.gbp.tt_sell": (50.0, 2000.0),
    "sl.fx.eur.tt_buy": (50.0, 2000.0),
    "sl.fx.eur.tt_sell": (50.0, 2000.0),
    "sl.fx.jpy.tt_buy": (0.1, 20.0),
    "sl.fx.jpy.tt_sell": (0.1, 20.0),
    "sl.fx.usd.ytd_change_pct": (-50.0, 50.0),
    "sl.fx.usd.week_avg_buy": (50.0, 1000.0),
    "sl.fx.usd.week_avg_sell": (50.0, 1000.0),
    "sl.fx.usd.week_avg_mid": (50.0, 1000.0),
    "sl.fx.usd.week_avg_mid_wow_pct": (-20.0, 20.0),
    "sl.fx.usd.week_avg_mid_yoy_pct": (-50.0, 50.0),
    "sl.fx.gbp.week_avg_buy": (50.0, 2000.0),
    "sl.fx.gbp.week_avg_sell": (50.0, 2000.0),
    "sl.fx.gbp.week_avg_mid": (50.0, 2000.0),
    "sl.fx.gbp.week_avg_mid_wow_pct": (-20.0, 20.0),
    "sl.fx.gbp.week_avg_mid_yoy_pct": (-50.0, 50.0),
    "sl.fx.eur.week_avg_buy": (50.0, 2000.0),
    "sl.fx.eur.week_avg_sell": (50.0, 2000.0),
    "sl.fx.eur.week_avg_mid": (50.0, 2000.0),
    "sl.fx.eur.week_avg_mid_wow_pct": (-20.0, 20.0),
    "sl.fx.eur.week_avg_mid_yoy_pct": (-50.0, 50.0),
    "sl.fx.jpy.week_avg_buy": (0.1, 20.0),
    "sl.fx.jpy.week_avg_sell": (0.1, 20.0),
    "sl.fx.jpy.week_avg_mid": (0.1, 20.0),
    "sl.fx.jpy.week_avg_mid_wow_pct": (-20.0, 20.0),
    "sl.fx.jpy.week_avg_mid_yoy_pct": (-50.0, 50.0),
    "sl.fx.usd.fwd_1m": (50.0, 1000.0),
    "sl.fx.usd.fwd_3m": (50.0, 1000.0),
    "sl.ei.currency_in_circulation": (100_000.0, 10_000_000.0),
    "sl.ei.reserve_money": (100_000.0, 10_000_000.0),
    "sl.eq.aspi": (100.0, 100_000.0),
    "sl.eq.sp_sl20": (100.0, 50_000.0),
    "sl.eq.turnover": (0.0, 500_000.0),
    "sl.eq.market_cap": (100.0, 100_000.0),
    "sl.eq.pe_ratio": (0.0, 100.0),
    "sl.eq.foreign_purchases": (0.0, 500_000.0),
    "sl.eq.foreign_sales": (0.0, 500_000.0),
    "sl.eq.foreign_net": (-500_000.0, 500_000.0),
    "sl.eq.aspi_wow_pct": (-30.0, 30.0),
    "sl.eq.sp_sl20_wow_pct": (-30.0, 30.0),
    "sl.ei.liquidity_surplus": (-5000.0, 5000.0),
    "sl.ei.remittances_usd": (0.0, 50_000.0),
    "sl.ei.remittances_usd_ytd": (0.0, 200_000.0),
    "sl.ei.tourist_arrivals": (0.0, 5_000_000.0),
    "sl.ei.tourist_arrivals_ytd": (0.0, 20_000_000.0),
    "sl.ei.tourist_earnings_usd": (0.0, 50_000.0),
    "sl.ei.tourist_earnings_usd_ytd": (0.0, 200_000.0),
    "sl.ei.pmi.manufacturing": (0.0, 100.0),
    "sl.ei.pmi.services": (0.0, 100.0),
    "sl.ei.pmi.construction": (0.0, 100.0),
    "sl.ei.m2b_yoy": (-50.0, 80.0),
    "sl.ei.credit.private": (100.0, 50_000.0),
    "sl.ei.credit.private_yoy": (-50.0, 80.0),
    "sl.ei.fiscal.primary_balance": (-50_000_000.0, 50_000_000.0),
    "sl.ei.fiscal.overall_balance": (-50_000_000.0, 50_000_000.0),
    "sl.ei.debt.domestic": (100.0, 100_000.0),
    "sl.ei.debt.foreign": (100.0, 100_000.0),
    "sl.ei.debt.total": (100.0, 200_000.0),
    "sl.ei.trade.exports_ytd": (0.0, 50_000.0),
    "sl.ei.trade.imports_ytd": (0.0, 80_000.0),
    "sl.ei.trade.balance_ytd": (-80_000.0, 50_000.0),
    "sl.ei.reserves.fwd_short": (-50_000.0, 0.0),
    "sl.ei.reserves.net_after_drains": (-50_000.0, 50_000.0),
    "sl.ei.electricity.thermal_oil": (0.0, 500.0),
    "sl.ei.electricity.thermal_oil_share": (0.0, 100.0),
    "sl.ei.energy.brent_wow": (-100.0, 100.0),
    "sl.fi.tbill.3m.bid_ask_spread": (0.0, 5.0),
    "sl.fi.tbond.5y.bid_ask_yield": (0.0, 5.0),
    "sl.ei.fuel.petrol_92": (50.0, 2000.0),
    "sl.ei.fuel.auto_diesel": (50.0, 2000.0),
    "sl.ei.fuel.kerosene": (20.0, 2000.0),
    "sl.ei.energy.brent": (10.0, 400.0),
    "sl.ei.energy.wti": (10.0, 400.0),
    "sl.ei.energy.opec": (10.0, 400.0),
    "sl.ei.energy.sg_petrol": (10.0, 400.0),
    "sl.ei.energy.sg_diesel": (10.0, 400.0),
    "sl.ei.energy.sg_kerosene": (10.0, 400.0),
    "sl.ei.electricity.generation": (0.0, 500.0),
    "sl.ei.electricity.peak_demand": (100.0, 10_000.0),
    "sl.fi.tbill.91d.primary": (0.0, 40.0),
    "sl.fi.tbill.182d.primary": (0.0, 40.0),
    "sl.fi.tbill.364d.primary": (0.0, 40.0),
    "sl.fi.tbill.91d.secondary": (0.0, 40.0),
    "sl.fi.tbill.182d.secondary": (0.0, 40.0),
    "sl.fi.tbill.364d.secondary": (0.0, 40.0),
    "sl.fi.tbond.2y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.3y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.4y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.5y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.6y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.8y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.10y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.15y.secondary_mid": (0.0, 40.0),
    "sl.fi.tbond.20y.secondary_mid": (0.0, 40.0),
    "sl.fi.isb.2028_04_pdi": (0.0, 40.0),
    "sl.fi.isb.2030_01_macro": (0.0, 40.0),
    "sl.fi.isb.2033_03_macro": (0.0, 40.0),
    "sl.fi.isb.2035_06_gov": (0.0, 40.0),
    "sl.fi.isb.2036_05_macro": (0.0, 40.0),
    "sl.fi.isb.2038_02_macro": (0.0, 40.0),
    "sl.fi.isb.2038_06_usd": (0.0, 40.0),
    "sl.fi.tbill.auction.offered": (0.0, 50_000_000.0),
    "sl.fi.tbill.auction.bids": (0.0, 50_000_000.0),
    "sl.fi.tbill.auction.accepted": (0.0, 50_000_000.0),
    "sl.fi.tbill.auction.cover": (0.0, 50.0),
    "sl.fi.gov.foreign_holdings": (0.0, 50_000_000.0),
    "sl.fi.gov.tbill_stock": (0.0, 50_000_000.0),
    "sl.fi.gov.tbond_stock": (0.0, 50_000_000.0),
    "sl.ei.gdp.growth": (-30.0, 30.0),
    "sl.ei.gdp.agriculture_yoy": (-40.0, 40.0),
    "sl.ei.gdp.industry_yoy": (-40.0, 40.0),
    "sl.ei.gdp.services_yoy": (-40.0, 40.0),
    "sl.ei.gdp.level": (100_000.0, 50_000_000.0),
    "sl.ei.iip": (10.0, 300.0),
    "sl.ei.fiscal.revenue_grants": (0.0, 50_000_000.0),
    "sl.ei.fiscal.recurrent_expenditure": (0.0, 50_000_000.0),
    "sl.ei.m1": (100.0, 50_000.0),
    "sl.ei.m2": (100.0, 100_000.0),
    "sl.ei.m2b": (100.0, 100_000.0),
    "sl.ei.m4": (100.0, 100_000.0),
    "sl.ei.trade.exports": (0.0, 10_000.0),
    "sl.ei.trade.imports": (0.0, 20_000.0),
    "sl.ei.trade.balance": (-20_000.0, 10_000.0),
    # MEI summary — agriculture / energy production (reference-month prints)
    "sl.ei.agri.tea": (0.0, 500.0),
    "sl.ei.agri.rubber": (0.0, 200.0),
    "sl.ei.agri.coconut": (0.0, 10_000.0),
    "sl.ei.agri.fish": (0.0, 1_000.0),
    "sl.ei.agri.paddy": (0.0, 20_000.0),
    "sl.ei.electricity.generation_monthly": (100.0, 20_000.0),
    # MEI summary — balance of payments (also written monthly by ESP)
    "sl.ei.bop.trade_account": (-50_000.0, 50_000.0),
    "sl.ei.bop.current_account": (-50_000.0, 50_000.0),
    "sl.ei.bop.current_account_ytd": (-50_000.0, 50_000.0),
    "sl.ei.bop.current_capital_account": (-50_000.0, 50_000.0),
    "sl.ei.bop.financial_account": (-50_000.0, 50_000.0),
    # Sri Lanka ORA has historically stayed well under ~10bn; MEI "Total Reserves" ~12bn must not pass
    "sl.ei.total_reserves": (100.0, 10_000.0),
    "sl.ei.reserves.fx": (50.0, 50_000.0),
    "sl.ei.reserves.imf": (0.0, 5_000.0),
    "sl.ei.reserves.sdrs": (0.0, 5_000.0),
    "sl.ei.reserves.gold": (0.0, 10_000.0),
    "sl.ei.reserves.other": (0.0, 5_000.0),
    "sl.ei.reserves.gor": (100.0, 20_000.0),
    "sl.ei.reserves.change": (-10_000.0, 10_000.0),
    "sl.ei.reserves.change_ytd": (-20_000.0, 20_000.0),
    # ESP services / income / financial flows
    "sl.ei.services.net": (-10_000.0, 10_000.0),
    "sl.ei.services.net_ytd": (-50_000.0, 50_000.0),
    "sl.ei.services.inflows": (0.0, 10_000.0),
    "sl.ei.services.inflows_ytd": (0.0, 50_000.0),
    "sl.ei.services.outflows": (0.0, 10_000.0),
    "sl.ei.services.outflows_ytd": (0.0, 50_000.0),
    "sl.ei.services.transport_inflows": (0.0, 5_000.0),
    "sl.ei.services.transport_inflows_ytd": (0.0, 20_000.0),
    "sl.ei.services.transport_outflows": (0.0, 5_000.0),
    "sl.ei.services.transport_outflows_ytd": (0.0, 20_000.0),
    "sl.ei.services.it_bpo": (0.0, 5_000.0),
    "sl.ei.services.it_bpo_ytd": (0.0, 20_000.0),
    "sl.ei.services.travel_abroad": (0.0, 5_000.0),
    "sl.ei.services.travel_abroad_ytd": (0.0, 20_000.0),
    "sl.ei.income.primary_net": (-10_000.0, 10_000.0),
    "sl.ei.income.primary_net_ytd": (-50_000.0, 50_000.0),
    "sl.ei.income.primary_inflows": (0.0, 5_000.0),
    "sl.ei.income.primary_inflows_ytd": (0.0, 20_000.0),
    "sl.ei.income.primary_outflows": (0.0, 5_000.0),
    "sl.ei.income.primary_outflows_ytd": (0.0, 20_000.0),
    "sl.ei.income.fdi_related": (0.0, 5_000.0),
    "sl.ei.income.fdi_related_ytd": (0.0, 20_000.0),
    "sl.ei.income.portfolio_related": (0.0, 5_000.0),
    "sl.ei.income.portfolio_related_ytd": (0.0, 20_000.0),
    "sl.ei.income.other_interest": (0.0, 5_000.0),
    "sl.ei.income.other_interest_ytd": (0.0, 20_000.0),
    "sl.ei.income.secondary_net": (-5_000.0, 10_000.0),
    "sl.ei.income.secondary_net_ytd": (-20_000.0, 50_000.0),
    "sl.ei.income.personal_outflows": (0.0, 5_000.0),
    "sl.ei.income.personal_outflows_ytd": (0.0, 20_000.0),
    "sl.ei.flows.cse": (-5_000.0, 5_000.0),
    "sl.ei.flows.cse_ytd": (-10_000.0, 10_000.0),
    "sl.ei.flows.gov_securities": (-5_000.0, 5_000.0),
    "sl.ei.flows.gov_securities_ytd": (-10_000.0, 10_000.0),
    "sl.ei.trade.exports_yoy": (-100.0, 200.0),
    "sl.ei.trade.imports_yoy": (-100.0, 200.0),
    "sl.ei.services.net_yoy": (-100.0, 200.0),
    "sl.ei.remittances_usd_yoy": (-100.0, 200.0),
    "sl.ei.tourist_earnings_usd_yoy": (-100.0, 200.0),
}

PRODUCTION_REPORTS = ("5206", "1059", "1064", "6277")
LOCKED_REPORTS = {"5206", "1059", "1064"}  # frozen mapping — force needs --unlock


def business_days_ago(n: int, end: date | None = None) -> date:
    d = end or date.today()
    left = n
    while left > 0:
        d -= timedelta(days=1)
        if d.weekday() < 5:
            left -= 1
    return d


def parse_number(text: str) -> float | None:
    t = (text or "").strip().replace(",", "")
    if not t or t in {"-", "—", "n/a", "N/A"}:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def find_data_table(soup: BeautifulSoup) -> Any:
    tables = soup.find_all("table")
    best = None
    best_score = 0
    for table in tables:
        rows = table.find_all("tr")
        text = table.get_text(" ", strip=True)
        date_hits = len(re.findall(r"\d{4}-\d{2}-\d{2}", text))
        score = date_hits * 10 + len(rows)
        if "DATA LIBRARY" in text and date_hits == 0:
            continue
        if score > best_score:
            best = table
            best_score = score
    return best


def extract_rows(html: str) -> list[list[str]]:
    soup = BeautifulSoup(html, "lxml")
    table = find_data_table(soup)
    if not table:
        return []
    rows: list[list[str]] = []
    for tr in table.find_all("tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if cells:
            rows.append(cells)
    return rows


def looks_like_date(value: str) -> bool:
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", value.strip()))


def obs(
    series_id: str,
    period: str,
    value: float,
    source_url: str,
    confidence: float,
) -> dict[str, Any] | None:
    bounds = BOUNDS.get(series_id)
    if bounds and not (bounds[0] <= value <= bounds[1]):
        return None
    return {
        "seriesId": series_id,
        "period": period,
        "value": value,
        "sourceUrl": source_url,
        "asOf": period,
        "confidence": round(confidence, 3),
    }


def normalize_mm_summary(rows: list[list[str]], source_url: str) -> list[dict[str, Any]]:
    """CBSL Money Market Summary (5206) — LOCKED mapping.

    Data row:
      [Date, CallMin, CallMax, CallWA, CallVol, RepoMin, RepoMax, RepoWA, RepoVol]
    """
    out: list[dict[str, Any]] = []
    mapping = [
        (1, "sl.mm.call.min_rate"),
        (2, "sl.mm.call.max_rate"),
        (3, "sl.mm.call.wa_yield"),
        (4, "sl.mm.call.volume"),
        (5, "sl.mm.repo.min_rate"),
        (6, "sl.mm.repo.max_rate"),
        (7, "sl.mm.repo.wa_yield"),
        (8, "sl.mm.repo.volume"),
    ]
    for row in rows:
        if not row or not looks_like_date(row[0]):
            continue
        period = row[0]
        for idx, series_id in mapping:
            if idx >= len(row):
                continue
            value = parse_number(row[idx])
            if value is None:
                continue
            item = obs(series_id, period, value, source_url, 0.95)
            if item:
                out.append(item)
    return out


def normalize_daily_ops(rows: list[list[str]], source_url: str) -> list[dict[str, Any]]:
    """CBSL Daily Operations (1059) — LOCKED mapping (Jul 2026).

    Data row:
      [Date, OfferRepo, OfferRev, Received, Accepted, Min, Max, WA,
       SDF_Vol, SDF_Rate, SLF_Vol, SLF_Rate, ON_Liquidity, Holdings]

    Exactly one of OfferRepo / OfferRev is filled per day (absorb vs inject).
    """
    out: list[dict[str, Any]] = []
    mapping = [
        (1, "sl.mm.omo.offer_repo"),
        (2, "sl.mm.omo.offer_reverse_repo"),
        (3, "sl.mm.omo.received"),
        (4, "sl.mm.omo.accepted"),
        (5, "sl.mm.omo.min_rate"),
        (6, "sl.mm.omo.max_rate"),
        (7, "sl.mm.omo.wa_yield"),
        (8, "sl.mm.sdf.volume"),
        (9, "sl.mm.sdf.rate"),
        (10, "sl.mm.slf.volume"),
        (11, "sl.mm.slf.rate"),
        (12, "sl.mm.overnight_liquidity"),
        (13, "sl.mm.cbsl.gov_holdings"),
    ]
    for row in rows:
        if not row or not looks_like_date(row[0]):
            continue
        if len(row) < 14:
            continue
        period = row[0]
        for idx, series_id in mapping:
            value = parse_number(row[idx])
            if value is None:
                continue
            item = obs(series_id, period, value, source_url, 0.93)
            if item:
                out.append(item)
    return out


def normalize_gov_sec(rows: list[list[str]], source_url: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    mapping = [
        (1, "sl.mm.tbill.91d"),
        (2, "sl.mm.tbill.182d"),
        (3, "sl.mm.tbill.364d"),
    ]
    for row in rows:
        if not row or not looks_like_date(row[0]):
            continue
        period = row[0]
        for idx, series_id in mapping:
            if idx >= len(row):
                continue
            value = parse_number(row[idx])
            if value is None:
                continue
            item = obs(series_id, period, value, source_url, 0.95)
            if item:
                out.append(item)
    return out


def date_to_yyyymmdd(value: str) -> float | None:
    if not looks_like_date(value):
        return None
    return float(value.replace("-", ""))


def normalize_term_repo(rows: list[list[str]], source_url: str) -> list[dict[str, Any]]:
    """CBSL OMO Term Repo / Reverse Repo (1064) — LOCKED mapping.

    Data row:
      [AuctionDate, SettlementDate, OfferRepo, OfferRev, Received, Accepted,
       Min, Max, WA, MaturityDate, TenureDays]

    Period = auction date. Full row retained so outstanding term book can be derived:
    settlement ≤ asOf < maturity, split by side (repo vs reverse repo).
    """
    out: list[dict[str, Any]] = []
    for row in rows:
        if not row or not looks_like_date(row[0]):
            continue
        if len(row) < 11:
            continue
        period = row[0]
        offer_repo = parse_number(row[2])
        offer_rr = parse_number(row[3])
        received = parse_number(row[4])
        accepted = parse_number(row[5])
        min_rate = parse_number(row[6])
        max_rate = parse_number(row[7])
        wa_yield = parse_number(row[8])
        tenure = parse_number(row[10])
        settlement = date_to_yyyymmdd(row[1])
        maturity = date_to_yyyymmdd(row[9])

        # Side: 1 = term repo (absorb), 2 = term reverse repo (inject)
        side: float | None = None
        if (offer_repo or 0) > 0 and (offer_rr or 0) <= 0:
            side = 1.0
        elif (offer_rr or 0) > 0 and (offer_repo or 0) <= 0:
            side = 2.0
        elif (offer_repo or 0) >= (offer_rr or 0) and (offer_repo or 0) > 0:
            side = 1.0
        elif (offer_rr or 0) > 0:
            side = 2.0

        fields: list[tuple[str, float | None]] = [
            ("sl.mm.term_repo.offer_repo", offer_repo),
            ("sl.mm.term_repo.offer_reverse_repo", offer_rr),
            ("sl.mm.term_repo.received", received),
            ("sl.mm.term_repo.accepted", accepted),
            ("sl.mm.term_repo.tenure", tenure),
            ("sl.mm.term_repo.settlement_date", settlement),
            ("sl.mm.term_repo.maturity_date", maturity),
            ("sl.mm.term_repo.side", side),
        ]
        # Rates only when auction filled — zeros on no-bid days are not real rates
        if (accepted or 0) > 0:
            fields.extend(
                [
                    ("sl.mm.term_repo.min_rate", min_rate),
                    ("sl.mm.term_repo.max_rate", max_rate),
                    ("sl.mm.term_repo.wa_yield", wa_yield),
                ]
            )

        for series_id, value in fields:
            if value is None:
                continue
            if series_id.endswith(("_rate", "wa_yield")) and value <= 0:
                continue
            item = obs(series_id, period, value, source_url, 0.94)
            if item:
                out.append(item)
    return out


def normalize_lending(rows: list[list[str]], source_url: str) -> list[dict[str, Any]]:
    """CBSL Commercial Bank Lending & Deposit Rates (6277).

    Data row (sparse — empty cells are normal):
      [EndWeek, AWLR, AWPR_Weekly, AWPR_Monthly, AWPR_6m,
       Deposit_Monthly, Deposit_6m, AWFDR, AWSR]
    """
    out: list[dict[str, Any]] = []
    mapping = [
        (1, "sl.mm.awlr"),
        (2, "sl.mm.awpr"),
        (3, "sl.mm.awpr.monthly"),
        (4, "sl.mm.awpr.6m"),
        (5, "sl.mm.awdr"),
        (6, "sl.mm.awdr.6m"),
        (7, "sl.mm.awfdr"),
        (8, "sl.mm.awsr"),
    ]
    for row in rows:
        if not row or not looks_like_date(row[0]):
            continue
        if len(row) < 3:
            continue
        period = row[0]
        for idx, series_id in mapping:
            if idx >= len(row):
                continue
            value = parse_number(row[idx])
            if value is None:
                continue
            item = obs(series_id, period, value, source_url, 0.92)
            if item:
                out.append(item)
    return out


NORMALIZERS = {
    "5206": normalize_mm_summary,
    "1059": normalize_daily_ops,
    "6169": normalize_gov_sec,
    "1064": normalize_term_repo,
    "6277": normalize_lending,
}


def parse_cbsl_change_date(value: Any) -> str | None:
    """Parse CBSL change-date cells like '27.11.2024' or '10.11.2004 (Close of Business)'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def normalize_policy_rates_xlsx(data: bytes, source_url: str) -> list[dict[str, Any]]:
    """CBSL historical_policy_interest_rates.xlsx — OPR + SRR sheets (step changes).

    Sheets are lightly formatted: leading blank columns, date as DD.MM.YYYY text,
    rate in the next numeric cell.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), data_only=True)
    out: list[dict[str, Any]] = []
    sheet_map = {
        "OPR": "sl.mm.opr",
        "SRR": "sl.mm.srr",
    }
    for sheet_name, series_id in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            cells = list(row)
            period = None
            date_idx = -1
            for i, cell in enumerate(cells):
                period = parse_cbsl_change_date(cell)
                if period:
                    date_idx = i
                    break
            if not period or date_idx < 0:
                continue
            value = None
            for cell in cells[date_idx + 1 :]:
                if cell is None or str(cell).strip() == "":
                    continue
                value = parse_number(str(cell))
                if value is not None:
                    break
            if value is None:
                continue
            item = obs(series_id, period, value, source_url, 0.98)
            if item:
                out.append(item)
    return out


def fetch_policy_rates_xlsx() -> tuple[str, bytes, str]:
    """Fetch policy-rates page, resolve historical xlsx link, download workbook."""
    src = SOURCES["policy-rates"]
    page_url = src["entry"]
    with httpx.Client(follow_redirects=True, timeout=120.0, headers=HEADERS) as client:
        page = client.get(page_url)
        page.raise_for_status()
        soup = BeautifulSoup(page.text, "lxml")
        xlsx_url = None
        for a in soup.find_all("a"):
            href = (a.get("href") or "").strip()
            if not href:
                continue
            if "historical_policy_interest_rates" in href.lower() or (
                href.lower().endswith(".xlsx") and "policy" in href.lower()
            ):
                xlsx_url = str(httpx.URL(str(page.url)).join(href))
                break
        if not xlsx_url:
            xlsx_url = src["xlsx_fallback"]
        xlsx = client.get(xlsx_url)
        xlsx.raise_for_status()
        if not xlsx.content:
            raise RuntimeError(f"Empty policy-rates workbook from {xlsx_url}")
        return str(page.url), xlsx.content, xlsx_url


MONTH_NAME_TO_NUM = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

CPI_SERIES_COLS = (
    (1, "sl.ei.ccpi.headline_yoy"),
    (2, "sl.ei.ccpi.core_yoy"),
    (3, "sl.ei.ncpi.headline_yoy"),
    (4, "sl.ei.ncpi.core_yoy"),
)


def normalize_cpi_html(html: str, source_url: str) -> list[dict[str, Any]]:
    """CBSL inflationwindow.php — monthly CCPI/NCPI headline + core YoY.

    Page embeds one or more tables (2021=100 current, 2013=100 history).
    Year rows (colspan) then month rows with 4 value columns.
    Prefer the first table when the same period appears twice.
    Period stored as month-start ISO date (YYYY-MM-01).
    """
    soup = BeautifulSoup(html, "lxml")
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    for table in soup.find_all("table"):
        text = table.get_text(" ", strip=True)
        if "Headline Inflation" not in text or "CCPI" not in text.upper():
            continue
        year: int | None = None
        for tr in table.find_all("tr"):
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
            if not cells:
                continue
            first = cells[0].strip()
            if re.fullmatch(r"\d{4}", first) and (
                len(cells) == 1 or all(not c.strip() or c.strip() == first for c in cells[1:])
            ):
                year = int(first)
                continue
            month = MONTH_NAME_TO_NUM.get(first.lower())
            if month is None or year is None or len(cells) < 5:
                continue
            period = f"{year:04d}-{month:02d}-01"
            for idx, series_id in CPI_SERIES_COLS:
                value = parse_number(cells[idx])
                if value is None:
                    continue
                key = (series_id, period)
                if key in seen:
                    continue
                item = obs(series_id, period, value, source_url, 0.96)
                if item:
                    seen.add(key)
                    out.append(item)
    return out


def fetch_cpi_table_html() -> tuple[str, str, str]:
    """Fetch CPI landing page + iframe table HTML."""
    src = SOURCES["consumer-price-inflation"]
    page_url = src["entry"]
    table_url = src["table"]
    with httpx.Client(follow_redirects=True, timeout=120.0, headers=HEADERS) as client:
        page = client.get(page_url)
        page.raise_for_status()
        table = client.get(table_url)
        table.raise_for_status()
        if not table.text.strip():
            raise RuntimeError(f"Empty CPI table from {table_url}")
        return str(page.url), table.text, str(table.url)


def fetch_report(report_id: str, from_date: date, to_date: date) -> tuple[str, str]:
    src = SOURCES[report_id]
    with httpx.Client(follow_redirects=True, timeout=120.0, headers=HEADERS) as client:
        r = client.get(src["entry"])
        r.raise_for_status()
        final_url = str(r.url)
        html = r.text
        if not html.strip():
            raise RuntimeError(f"Empty response from {final_url}")

        soup = BeautifulSoup(html, "lxml")
        form = soup.find("form")
        if not form:
            return final_url, html

        payload: dict[str, str] = {}
        for inp in form.find_all("input"):
            name = inp.get("name")
            if not name:
                continue
            if name.endswith("btnExport"):
                continue
            payload[name] = inp.get("value") or ""

        for key in list(payload.keys()):
            if key.endswith("txtFrom") or key.endswith("$txtFrom"):
                payload[key] = from_date.isoformat()
            if key.endswith("txtTo") or key.endswith("$txtTo"):
                payload[key] = to_date.isoformat()

        show_key = next((k for k in payload if k.endswith("btnShow")), None)
        if show_key:
            payload[show_key] = "Show"

        action = form.get("action") or final_url
        post_url = httpx.URL(final_url).join(action)
        post = client.post(str(post_url), data=payload)
        post.raise_for_status()
        if post.text.strip():
            html = post.text
            final_url = str(post.url)

    return final_url, html


def content_hash(html: str) -> str:
    return hashlib.sha256(html.encode("utf-8")).hexdigest()


def post_ingest(
    observations: list[dict[str, Any]],
    meta: dict[str, Any],
    dry_run: bool,
) -> dict[str, Any]:
    body = {"observations": observations, "document": meta}
    if dry_run:
        return {
            "mode": "dry_run",
            "count": len(observations),
            "sample": observations[:3],
            "document": {k: meta[k] for k in ("sourceReportId", "contentHash", "rawPath") if k in meta},
        }

    with httpx.Client(timeout=30.0) as client:
        r = client.post(f"{API_URL}/ingest/observations", json=body)
        r.raise_for_status()
        return r.json()


def parse_report(report_id: str, html: str, source_url: str) -> list[dict[str, Any]]:
    if report_id not in NORMALIZERS:
        raise ValueError(f"No normalizer for {report_id}")
    rows = extract_rows(html)
    if not rows:
        raise RuntimeError(f"No data table found for report {report_id}")
    return NORMALIZERS[report_id](rows, source_url)


def run_self_test() -> int:
    """Parse fixtures and assert expected series ids appear."""
    expected = {
        "5206": {
            "sl.mm.call.min_rate",
            "sl.mm.call.max_rate",
            "sl.mm.call.wa_yield",
            "sl.mm.call.volume",
            "sl.mm.repo.min_rate",
            "sl.mm.repo.max_rate",
            "sl.mm.repo.wa_yield",
            "sl.mm.repo.volume",
        },
        "1059": {
            "sl.mm.omo.offer_repo",
            "sl.mm.omo.received",
            "sl.mm.omo.accepted",
            "sl.mm.omo.min_rate",
            "sl.mm.omo.max_rate",
            "sl.mm.omo.wa_yield",
            "sl.mm.sdf.volume",
            "sl.mm.sdf.rate",
            "sl.mm.slf.volume",
            "sl.mm.slf.rate",
            "sl.mm.overnight_liquidity",
            "sl.mm.cbsl.gov_holdings",
        },
        "6169": {"sl.mm.tbill.91d", "sl.mm.tbill.182d", "sl.mm.tbill.364d"},
        "1064": {
            "sl.mm.term_repo.offer_repo",
            "sl.mm.term_repo.offer_reverse_repo",
            "sl.mm.term_repo.received",
            "sl.mm.term_repo.accepted",
            "sl.mm.term_repo.min_rate",
            "sl.mm.term_repo.max_rate",
            "sl.mm.term_repo.wa_yield",
            "sl.mm.term_repo.tenure",
            "sl.mm.term_repo.settlement_date",
            "sl.mm.term_repo.maturity_date",
            "sl.mm.term_repo.side",
        },
        "6277": {
            "sl.mm.awlr",
            "sl.mm.awpr",
            "sl.mm.awpr.monthly",
            "sl.mm.awpr.6m",
            "sl.mm.awdr",
            "sl.mm.awdr.6m",
            "sl.mm.awfdr",
            "sl.mm.awsr",
        },
    }
    failures = 0
    for report_id, need in expected.items():
        path = FIXTURES_DIR / f"{report_id}.html"
        if not path.exists():
            print(f"FAIL missing fixture {path}", file=sys.stderr)
            failures += 1
            continue
        html = path.read_text(encoding="utf-8")
        obs_list = parse_report(report_id, html, f"fixture://{report_id}")
        got = {o["seriesId"] for o in obs_list}
        missing = need - got
        if missing:
            print(f"FAIL {report_id} missing {missing}; got {sorted(got)}", file=sys.stderr)
            failures += 1
        else:
            print(f"OK {report_id}: {len(obs_list)} observations, series={sorted(got)}")

    cpi_path = FIXTURES_DIR / "consumer-price-inflation.html"
    cpi_need = {
        "sl.ei.ccpi.headline_yoy",
        "sl.ei.ccpi.core_yoy",
        "sl.ei.ncpi.headline_yoy",
        "sl.ei.ncpi.core_yoy",
    }
    if not cpi_path.exists():
        print(f"FAIL missing fixture {cpi_path}", file=sys.stderr)
        failures += 1
    else:
        cpi_obs = normalize_cpi_html(
            cpi_path.read_text(encoding="utf-8"),
            "fixture://consumer-price-inflation",
        )
        cpi_got = {o["seriesId"] for o in cpi_obs}
        cpi_missing = cpi_need - cpi_got
        # Spot-check a known print from the 2021=100 table.
        june_2026 = next(
            (
                o
                for o in cpi_obs
                if o["seriesId"] == "sl.ei.ccpi.headline_yoy" and o["period"] == "2026-06-01"
            ),
            None,
        )
        if cpi_missing:
            print(
                f"FAIL consumer-price-inflation missing {cpi_missing}; got {sorted(cpi_got)}",
                file=sys.stderr,
            )
            failures += 1
        elif not june_2026 or june_2026["value"] != 6.8:
            print(
                f"FAIL consumer-price-inflation expected CCPI headline 2026-06-01=6.8, got {june_2026}",
                file=sys.stderr,
            )
            failures += 1
        else:
            print(
                f"OK consumer-price-inflation: {len(cpi_obs)} observations, "
                f"series={sorted(cpi_got)}"
            )

    from economic_indicators import run_pdf_self_test

    for msg in run_pdf_self_test(FIXTURES_DIR, BOUNDS):
        print(f"FAIL {msg}", file=sys.stderr)
        failures += 1
    return 1 if failures else 0


def calendar_days_ago(n: int, end: date | None = None) -> date:
    d = end or date.today()
    return d - timedelta(days=n)


def iter_date_chunks(from_d: date, to_d: date, chunk_days: int) -> list[tuple[date, date]]:
    """Split [from_d, to_d] into inclusive windows of at most chunk_days calendar days."""
    if chunk_days < 1:
        raise ValueError("chunk_days must be >= 1")
    chunks: list[tuple[date, date]] = []
    cur = from_d
    while cur <= to_d:
        end = min(cur + timedelta(days=chunk_days - 1), to_d)
        chunks.append((cur, end))
        cur = end + timedelta(days=1)
    return chunks


def iter_month_chunks(from_d: date, to_d: date) -> list[tuple[date, date]]:
    """Split range into calendar-month windows (gentle CBSL pacing)."""
    chunks: list[tuple[date, date]] = []
    y, m = from_d.year, from_d.month
    while True:
        month_start = date(y, m, 1)
        if m == 12:
            month_end = date(y, 12, 31)
            next_y, next_m = y + 1, 1
        else:
            month_end = date(y, m + 1, 1) - timedelta(days=1)
            next_y, next_m = y, m + 1
        start = max(month_start, from_d)
        end = min(month_end, to_d)
        if start <= end:
            chunks.append((start, end))
        if date(next_y, next_m, 1) > to_d:
            break
        y, m = next_y, next_m
    return chunks


def fetch_and_ingest_policy_rates(*, dry_run: bool, force: bool) -> dict[str, Any]:
    page_url, xlsx_bytes, xlsx_url = fetch_policy_rates_xlsx()
    digest = hashlib.sha256(xlsx_bytes).hexdigest()
    raw_path = RAW_DIR / f"policy-rates_{digest[:10]}.xlsx"
    raw_path.write_bytes(xlsx_bytes)
    observations = normalize_policy_rates_xlsx(xlsx_bytes, xlsx_url)
    if not observations:
        return {
            "ok": True,
            "skipped": True,
            "observations": 0,
            "seriesIds": [],
            "rawPath": str(raw_path),
            "contentHash": digest,
            "htmlBytes": len(xlsx_bytes),
            "window": {"from": None, "to": None},
            "ingest": {"mode": "empty_window"},
        }
    periods = sorted({o["period"] for o in observations})
    meta = {
        "sourceReportId": "policy-rates",
        "url": page_url,
        "contentHash": digest,
        "fromDate": periods[0],
        "toDate": periods[-1],
        "rawPath": str(raw_path),
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "force": force,
    }
    ingest_result = post_ingest(observations, meta, dry_run=dry_run)
    return {
        "ok": True,
        "observations": len(observations),
        "seriesIds": sorted({o["seriesId"] for o in observations}),
        "rawPath": str(raw_path),
        "contentHash": digest,
        "htmlBytes": len(xlsx_bytes),
        "window": {"from": periods[0], "to": periods[-1]},
        "ingest": ingest_result,
        "xlsxUrl": xlsx_url,
    }


def fetch_and_ingest_cpi(*, dry_run: bool, force: bool) -> dict[str, Any]:
    page_url, html, table_url = fetch_cpi_table_html()
    digest = content_hash(html)
    raw_path = RAW_DIR / f"consumer-price-inflation_{digest[:10]}.html"
    raw_path.write_text(html, encoding="utf-8")
    observations = normalize_cpi_html(html, table_url)
    if not observations:
        return {
            "ok": True,
            "skipped": True,
            "observations": 0,
            "seriesIds": [],
            "rawPath": str(raw_path),
            "contentHash": digest,
            "htmlBytes": len(html),
            "window": {"from": None, "to": None},
            "ingest": {"mode": "empty_window"},
        }
    periods = sorted({o["period"] for o in observations})
    meta = {
        "sourceReportId": "consumer-price-inflation",
        "url": page_url,
        "contentHash": digest,
        "fromDate": periods[0],
        "toDate": periods[-1],
        "rawPath": str(raw_path),
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "force": force,
    }
    ingest_result = post_ingest(observations, meta, dry_run=dry_run)
    return {
        "ok": True,
        "observations": len(observations),
        "seriesIds": sorted({o["seriesId"] for o in observations}),
        "rawPath": str(raw_path),
        "contentHash": digest,
        "htmlBytes": len(html),
        "window": {"from": periods[0], "to": periods[-1]},
        "ingest": ingest_result,
        "tableUrl": table_url,
    }


def fetch_and_ingest_window(
    report_id: str,
    from_d: date,
    to_d: date,
    *,
    dry_run: bool,
    force: bool,
    delay_sec: float = 4.0,
) -> dict[str, Any]:
    if report_id == "policy-rates":
        return fetch_and_ingest_policy_rates(dry_run=dry_run, force=force)
    if report_id == "consumer-price-inflation":
        return fetch_and_ingest_cpi(dry_run=dry_run, force=force)
    if report_id in PDF_LISTING_REPORTS:
        from economic_indicators import fetch_and_ingest_pdf_reports

        return fetch_and_ingest_pdf_reports(
            report_id,
            from_d,
            to_d,
            dry_run=dry_run,
            force=force,
            raw_dir=RAW_DIR,
            bounds=BOUNDS,
            post_ingest=post_ingest,
            delay_sec=delay_sec,
            max_listing_pages=80,
        )

    final_url, html = fetch_report(report_id, from_d, to_d)
    digest = content_hash(html)
    raw_path = RAW_DIR / f"{report_id}_{from_d.isoformat()}_{to_d.isoformat()}_{digest[:10]}.html"
    raw_path.write_text(html, encoding="utf-8")
    try:
        observations = parse_report(report_id, html, final_url)
    except RuntimeError as exc:
        # Empty months / "No records found" — skip, don't fail the whole backfill
        if "No data table" in str(exc) or "zero observations" in str(exc).lower():
            return {
                "ok": True,
                "skipped": True,
                "observations": 0,
                "seriesIds": [],
                "rawPath": str(raw_path),
                "contentHash": digest,
                "htmlBytes": len(html),
                "window": {"from": from_d.isoformat(), "to": to_d.isoformat()},
                "ingest": {"mode": "empty_window"},
            }
        raise
    if not observations:
        return {
            "ok": True,
            "skipped": True,
            "observations": 0,
            "seriesIds": [],
            "rawPath": str(raw_path),
            "contentHash": digest,
            "htmlBytes": len(html),
            "window": {"from": from_d.isoformat(), "to": to_d.isoformat()},
            "ingest": {"mode": "empty_window"},
        }

    meta = {
        "sourceReportId": report_id,
        "url": final_url,
        "contentHash": digest,
        "fromDate": from_d.isoformat(),
        "toDate": to_d.isoformat(),
        "rawPath": str(raw_path),
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "force": force,
    }
    ingest_result = post_ingest(observations, meta, dry_run=dry_run)
    return {
        "ok": True,
        "observations": len(observations),
        "seriesIds": sorted({o["seriesId"] for o in observations}),
        "rawPath": str(raw_path),
        "contentHash": digest,
        "htmlBytes": len(html),
        "window": {"from": from_d.isoformat(), "to": to_d.isoformat()},
        "ingest": ingest_result,
    }


def run(
    report_ids: list[str],
    days: int,
    dry_run: bool,
    force: bool,
    unlock: bool,
    *,
    years: float | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    chunk_days: int = 120,
    by_month: bool = False,
    delay_sec: float = 4.0,
) -> int:
    to_d = to_date or date.today()
    if from_date:
        from_d = from_date
    elif years is not None and years > 0:
        from_d = calendar_days_ago(int(years * 365.25), to_d)
    else:
        from_d = business_days_ago(days, to_d)

    if from_d > to_d:
        raise ValueError(f"from_date {from_d} is after to_date {to_d}")

    chunks = (
        iter_month_chunks(from_d, to_d)
        if by_month
        else iter_date_chunks(from_d, to_d, chunk_days)
    )
    mode = "month" if by_month else f"≤{chunk_days}d"
    print(
        f"Ingest window {from_d} → {to_d} "
        f"({len(chunks)} chunk(s) × {mode}, delay {delay_sec}s)"
    )

    results: list[dict[str, Any]] = []
    total = 0
    any_fail = False
    fetch_i = 0

    for report_id in report_ids:
        if report_id not in SOURCES:
            results.append({"reportId": report_id, "ok": False, "error": "unknown_report"})
            any_fail = True
            continue

        if report_id in LOCKED_REPORTS and force and not unlock:
            msg = (
                f"Report {report_id} is locked (verified). "
                "Pass --unlock with --force to re-ingest intentionally."
            )
            print(f"  SKIP {report_id}: {msg}")
            results.append(
                {
                    "reportId": report_id,
                    "title": SOURCES[report_id]["title"],
                    "ok": True,
                    "skipped": True,
                    "error": "locked_force_needs_unlock",
                }
            )
            continue

        entry: dict[str, Any] = {
            "reportId": report_id,
            "title": SOURCES[report_id]["title"],
            "ok": True,
            "chunks": [],
            "observations": 0,
            "seriesIds": [],
        }
        print(f"Fetching {report_id} ({SOURCES[report_id]['title']})...")
        series_acc: set[str] = set()

        try:
            # Full-history / PDF-listing sources — one logical fetch over the window.
            report_chunks = (
                [(from_d, to_d)]
                if report_id in FULL_HISTORY_REPORTS or report_id in PDF_LISTING_REPORTS
                else chunks
            )
            for chunk_from, chunk_to in report_chunks:
                if fetch_i > 0 and delay_sec > 0 and report_id not in PDF_LISTING_REPORTS:
                    print(f"  … waiting {delay_sec:.1f}s (be kind to CBSL)")
                    time.sleep(delay_sec)
                fetch_i += 1
                if report_id == "policy-rates":
                    print("  workbook (full history)")
                elif report_id == "consumer-price-inflation":
                    print("  CPI table iframe (full history)")
                elif report_id in PDF_LISTING_REPORTS:
                    print(f"  PDF listing {chunk_from} → {chunk_to}")
                else:
                    print(f"  chunk {chunk_from} → {chunk_to}")
                chunk_result = fetch_and_ingest_window(
                    report_id,
                    chunk_from,
                    chunk_to,
                    dry_run=dry_run,
                    force=force,
                    delay_sec=delay_sec,
                )
                entry["chunks"].append(chunk_result)
                if chunk_result.get("ok") is False:
                    any_fail = True
                    entry["ok"] = False
                    entry["error"] = chunk_result.get("error") or "pdf_batch_failed"
                elif chunk_result.get("warning"):
                    entry["warning"] = chunk_result["warning"]
                total += int(chunk_result.get("observations") or 0)
                entry["observations"] += int(chunk_result.get("observations") or 0)
                series_acc.update(chunk_result.get("seriesIds") or [])
                if chunk_result.get("skipped"):
                    print("    empty (no rows) — skipped")
                elif report_id in PDF_LISTING_REPORTS:
                    fetched = chunk_result.get("fetched", "?")
                    ok_n = chunk_result.get("okFiles", "?")
                    fail_n = chunk_result.get("failedFiles", 0)
                    warn = chunk_result.get("warning")
                    print(
                        f"    OK → {chunk_result.get('observations', 0)} observations "
                        f"from {ok_n}/{fetched} PDF(s)"
                        + (f" ({warn})" if warn else "")
                    )
                    if fail_n:
                        print(f"    WARN {fail_n} PDF(s) skipped — see chunk errors")
                else:
                    print(f"    OK → {chunk_result['observations']} observations")
            entry["seriesIds"] = sorted(series_acc)
            print(
                f"  DONE {report_id} → {entry['observations']} observations "
                f"({', '.join(entry['seriesIds'])})"
            )
        except Exception as exc:  # noqa: BLE001
            any_fail = True
            entry["ok"] = False
            entry["error"] = str(exc)
            print(f"  FAIL {report_id}: {exc}", file=sys.stderr)

        results.append(entry)

    summary = {
        "ok": not any_fail,
        "dryRun": dry_run,
        "force": force,
        "unlock": unlock,
        "chunkDays": chunk_days,
        "delaySec": delay_sec,
        "window": {"from": from_d.isoformat(), "to": to_d.isoformat()},
        "totalObservations": total,
        "reports": results,
    }
    print("---SUMMARY_JSON---")
    print(json.dumps(summary, indent=2))
    print("---END_SUMMARY---")
    print(f"Done. {total} observations processed.")
    return 1 if any_fail else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="CBSL Money Market extractor (production slice)")
    parser.add_argument("--reports", default=",".join(PRODUCTION_REPORTS))
    parser.add_argument("--days", type=int, default=14, help="Business-day lookback (ignored if --years/--from set)")
    parser.add_argument(
        "--years",
        type=float,
        default=None,
        help="Calendar-year lookback (e.g. 5). Prefer this for long backfills.",
    )
    parser.add_argument(
        "--from",
        dest="from_date",
        default=None,
        help="Start date YYYY-MM-DD (overrides --days/--years)",
    )
    parser.add_argument(
        "--to",
        dest="to_date",
        default=None,
        help="End date YYYY-MM-DD (default: today)",
    )
    parser.add_argument(
        "--chunk-days",
        type=int,
        default=120,
        help="Max calendar days per CBSL request (default 120). Ignored with --by-month.",
    )
    parser.add_argument(
        "--by-month",
        action="store_true",
        help="Fetch one calendar month per request (gentlest for old history)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=4.0,
        help="Seconds to wait between CBSL requests (default 4).",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-ingest even if content hash already seen",
    )
    parser.add_argument(
        "--unlock",
        action="store_true",
        help="Required with --force for locked reports (5206, 1059, 1064)",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Parse local fixtures and exit",
    )
    args = parser.parse_args()
    if args.self_test:
        return run_self_test()
    report_ids = [x.strip() for x in args.reports.split(",") if x.strip()]

    def parse_iso(value: str | None) -> date | None:
        if not value:
            return None
        return date.fromisoformat(value)

    return run(
        report_ids,
        args.days,
        args.dry_run,
        args.force,
        args.unlock,
        years=args.years,
        from_date=parse_iso(args.from_date),
        to_date=parse_iso(args.to_date),
        chunk_days=args.chunk_days,
        by_month=args.by_month,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
